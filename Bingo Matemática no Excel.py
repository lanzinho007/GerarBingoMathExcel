import random
from openpyxl import Workbook

# Função para gerar uma cartela de bingo
def gerar_cartela():
  numeros = []
  for i in range(2, 6):
    for j in range(1, 11):
      numeros.append(i * j)
  random.shuffle(numeros)
  return numeros[:25]

# Função para gerar o arquivo Excel com as cartelas
def gerar_excel(cartelas):
  wb = Workbook()
  ws = wb.active

  # Cabeçalho
  ws["A1"] = "Cartela de Bingo"
  ws["A2"] = "Tabuadas: 2, 3, 4 e 5"

  # Gerar cartelas
  for i, cartela in enumerate(cartelas, start=1):
    linha_inicial = 4 + (i - 1) * 10
    for j in range(5):
      for k in range(5):
        ws.cell(row=linha_inicial + j, column=1 + k, value=cartela[j * 5 + k])

  # Ajustar formatação
  for i in range(1, 6):
    ws.column_dimensions[chr(ord("A") + i - 1)].width = 10
  ws.row_dimensions[1].height = 20

  # Salvar arquivo
  wb.save("cartelas_bingo.xlsx")

# Quantidade de cartelas
quantidade_cartelas = int(input("Digite a quantidade de cartelas desejadas: "))

# Gerar cartelas
cartelas = []
for i in range(quantidade_cartelas):
  cartelas.append(gerar_cartela())

# Gerar arquivo Excel
gerar_excel(cartelas)

print("As cartelas foram geradas no arquivo 'cartelas_bingo.xlsx'.")
